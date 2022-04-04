import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import NoSuchElementException

print('Логин: ')
login = input()
print('Пароль: ')
password = input()

# Настройка профиля Мозиллы(не спрашивает о сохранении, скачивает в указанную нами директорию)
fp = webdriver.FirefoxProfile()
fp.set_preference("browser.download.folderList",2)
fp.set_preference("browser.download.manager.showWhenStarting", False)

driver = webdriver.Firefox(firefox_profile=fp)
wait = WebDriverWait(driver, 60)

#connection refused, permission denied
driver.get('https://accounts.vetrf.ru/')
time.sleep(5)

#Вход в систему
element = driver.find_element_by_xpath("/html/body/div[2]/div[2]/div[1]/button")
element.click()
time.sleep(5)

#Логин\пароль\вход
element = wait.until(ec.presence_of_element_located((By.XPATH, '//*[@id="username"]')))
element.send_keys(login)

element = driver.find_element_by_xpath('//*[@id="password"]')
element.send_keys(password)

element = driver.find_element_by_xpath('/html/body/div[2]/div/div[2]/center[1]/div/form/div[5]/button').click()

#Список пользователей
element = wait.until(ec.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/ul/li[2]/a')))
element.click()

time.sleep(15)

wb = load_workbook('sotrudniki.xlsx')
ws = wb['Лист1']

for i in range(1, ws.max_row+1):
    if ws.cell(row=i, column=1).value is not None:

        #Находим поле и вставляем в него фио, жмем ENTER
        element = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div[2]/div[2]/div/div[3]/table/thead/tr[2]/th[1]/input')
        element.send_keys(ws.cell(row=i, column=1).value)
        element.send_keys(Keys.RETURN)

        time.sleep(5)

        #Клик на профиль
        try:
            element = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr/td[5]/div/a[1]/i')
            element.click()

            time.sleep(3)
        except NoSuchElementException:
            element = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div[2]/div[2]/div/div[3]/table/thead/tr[2]/th[1]/input')
            element.send_keys(Keys.CONTROL + 'a')
            element.send_keys(Keys.DELETE)
            print(ws.cell(row=i, column=1).value, ' Не заведен')
            continue

        #Прокрутка к хоз.субъектам и клик на редактирование
        element = wait.until(ec.element_to_be_clickable((By.XPATH,'/html/body/div[2]/div[2]/div/div[2]/div[2]/div/div/div/div[1]/div/div/div/h4[5]/i')))
        element.click()
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(3)

        #Цикл добавления субъектов, если уже все добавлено - клик на сохранение
        while True:
            try:
                element = driver.find_element_by_xpath('//*[@id="profile_edit_businessEntities_chosen"]').click()
                time.sleep(1)
                element = driver.find_element_by_xpath('//*[contains(@class, "active-result")]').click()
                time.sleep(1)
                element = driver.find_element_by_xpath('//*[contains(@title, "Добавить")]')
                element.click()
                time.sleep(1)
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(1)
            except ElementNotInteractableException:
                break
            else:
                continue
        element = driver.find_element_by_xpath('//*[contains(@name, "businessEntitiesSaveButton")]')
        element.click()
        time.sleep(45)

        #Прокрутка к хоз.субъектам и клик на редактирование их
        element = driver.find_element_by_xpath('//*[contains(@name, "businessEntities")]')
        driver.execute_script("arguments[0].scrollIntoView();", element)
        time.sleep(1)
        element.click()

        time.sleep(1)

        #Циикл: клик на редактирование субъектов, клик на сам объект и проставление галочек Гашение ВСД, Возвратные ВСД
        #  и приказ 6***, если галки стоят - проверяет и пропускает
        vse = driver.find_elements_by_xpath('//*[contains(@name, "deleteBE_btn_")][@class="tooltip-info my-tooltip-link"]')
        n = 1
        while n <= len(vse):
            try:
                vse = driver.find_elements_by_xpath('//*[contains(@name, "deleteBE_btn_")][@class="tooltip-info my-tooltip-link"]')
                driver.execute_script("arguments[0].scrollIntoView();", vse[n-1])
                time.sleep(1)
                vse[n-1].click()
                time.sleep(2)
            except ElementNotInteractableException:
                continue
            time.sleep(2)
            gashenie = driver.find_element_by_xpath('//*[@id="external_auth_3"]')
            if gashenie.is_selected() is True:
                pass
            else:
                gashenie.click()
            vozvrat = driver.find_element_by_xpath('//*[@id="external_auth_11"]')
            if vozvrat.is_selected() is True:
                pass
            else:
                vozvrat.click()
            shest = driver.find_element_by_xpath('//*[@id="external_auth_14"]')
            if shest.is_selected() is True:
                pass
            else:
                shest.click()
            time.sleep(1)
            element = driver.find_element_by_xpath('//*[contains(@type, "submit")]')
            element.click()
            time.sleep(2)
            element = driver.find_element_by_xpath('//*[contains(@name, "businessEntities")]')
            driver.execute_script("arguments[0].scrollIntoView();", element)
            element.click()
            time.sleep(1)
            n += 1
        print(ws.cell(row=i, column=1).value, ' Готово')
        time.sleep(4)
        element = driver.find_element_by_xpath('/html/body/div[2]/div[1]/ul/li[2]/a').click()
        time.sleep(20)


time.sleep(3)

#Завершение работы бота и закрытие браузера
print('Завершено')
driver.quit()